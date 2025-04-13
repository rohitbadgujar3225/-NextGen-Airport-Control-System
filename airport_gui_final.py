import tkinter as tk
from tkinter import ttk, messagebox
import mysql.connector
from mysql.connector import Error
import json
import logging

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='airport_gui.log'
)
logger = logging.getLogger(__name__)

class AirportDBGUI:
    def __init__(self, root):
        logger.debug("Initializing AirportDBGUI")
        self.root = root
        self.root.title("✈ Airport Database Management System")
        logger.debug("GUI window created")
        self.root.geometry("800x600")
        
        # Add watermark
        self.watermark = tk.Label(self.root, text="AIRPORT DBMS", 
                                font=("Arial", 72), fg="#f0f0f0")
        self.watermark.place(relx=0.5, rely=0.5, anchor="center")
        self.watermark.lower()  # Send to back
        
        # Initialize connection and fields
        self.connection = None
        self.add_entries = {}
        self.db_config = {
            'host': 'localhost',
            'database': 'airportmanagementsystem',
            'user': 'root',
            'password': '3225'  # Added the correct password
        }
        
        self.create_widgets()
        if not self.connect_to_db():
            logger.error("Failed to connect to database")
            messagebox.showerror("Error", "Failed to connect to database")
        else:
            logger.info("Database connection successful")

    def create_widgets(self):
        # Main container frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Left panel for operations
        left_panel = ttk.Frame(main_frame)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        
        # Right panel for results
        right_panel = ttk.Frame(main_frame)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Query section
        query_frame = ttk.LabelFrame(left_panel, text="SQL Query", padding="10")
        query_frame.pack(fill=tk.X, pady=5)
        
        self.query_entry = tk.Text(query_frame, height=4, wrap=tk.WORD)
        self.query_entry.pack(fill=tk.X, pady=5)
        
        ttk.Button(query_frame, text="Execute", command=self.execute_query).pack(pady=5)
        
        # Quick query buttons
        quick_frame = ttk.LabelFrame(left_panel, text="Quick Operations", padding="10")
        quick_frame.pack(fill=tk.X, pady=5)
        
        # Flight operations
        flight_frame = ttk.LabelFrame(quick_frame, text="Flights")
        flight_frame.pack(fill=tk.X, pady=2)
        
        ttk.Button(flight_frame, text="View All", 
                 command=lambda: self.set_query("SELECT * FROM flight")).pack(side=tk.LEFT, expand=True)
        ttk.Button(flight_frame, text="Active Flights", 
                 command=lambda: self.set_query("SELECT * FROM flight WHERE departure_time > NOW()")).pack(side=tk.LEFT, expand=True)
        ttk.Button(flight_frame, text="Flight Count", 
                 command=lambda: self.set_query("SELECT COUNT(*) AS total_flights FROM flight")).pack(side=tk.LEFT, expand=True)
        
        # Passenger operations
        passenger_frame = ttk.LabelFrame(quick_frame, text="Passengers")
        passenger_frame.pack(fill=tk.X, pady=2)
        
        ttk.Button(passenger_frame, text="View All", 
                 command=lambda: self.set_query("SELECT * FROM passenger")).pack(side=tk.LEFT, expand=True)
        ttk.Button(passenger_frame, text="VIP Passengers", 
                 command=lambda: self.set_query("SELECT * FROM passenger WHERE vip_status = 1")).pack(side=tk.LEFT, expand=True)
        ttk.Button(passenger_frame, text="Passenger Count", 
                 command=lambda: self.set_query("SELECT COUNT(*) AS total_passengers FROM passenger")).pack(side=tk.LEFT, expand=True)

        # Schedule operations
        schedule_frame = ttk.LabelFrame(quick_frame, text="Schedules")
        schedule_frame.pack(fill=tk.X, pady=2)
        
        ttk.Button(schedule_frame, text="View All", 
                 command=lambda: self.set_query("SELECT * FROM schedule")).pack(side=tk.LEFT, expand=True)
        ttk.Button(schedule_frame, text="Upcoming", 
                 command=lambda: self.set_query("SELECT * FROM schedule WHERE departure_time > NOW()")).pack(side=tk.LEFT, expand=True)

        # Journey operations
        journey_frame = ttk.LabelFrame(quick_frame, text="Journeys")
        journey_frame.pack(fill=tk.X, pady=2)
        
        ttk.Button(journey_frame, text="View All", 
                 command=lambda: self.set_query("SELECT * FROM journey")).pack(side=tk.LEFT, expand=True)

        # Ticket operations
        ticket_frame = ttk.LabelFrame(quick_frame, text="Tickets")
        ticket_frame.pack(fill=tk.X, pady=2)
        
        ttk.Button(ticket_frame, text="View All", 
                 command=lambda: self.set_query("SELECT * FROM ticket")).pack(side=tk.LEFT, expand=True)

        # Details operations
        details_frame = ttk.LabelFrame(quick_frame, text="Details")
        details_frame.pack(fill=tk.X, pady=2)

        # Create menu button
        self.details_btn = ttk.Menubutton(details_frame, text="Add Details")
        self.details_menu = tk.Menu(self.details_btn, tearoff=0)
        self.details_btn["menu"] = self.details_menu
        
        self.details_menu.add_command(label="Add New Record", command=self.show_add_dialog)
        self.details_menu.add_command(label="Delete Selected", command=self.delete_selected)
        self.details_btn.pack(side=tk.LEFT, expand=True)
        
        
        # Results section
        results_frame = ttk.LabelFrame(right_panel, text="Results", padding="10")
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Filter controls
        filter_frame = ttk.Frame(results_frame)
        filter_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(filter_frame, text="Filter by:").pack(side=tk.LEFT, padx=5)
        self.filter_column = ttk.Combobox(filter_frame, state="readonly")
        self.filter_column.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        ttk.Label(filter_frame, text="Value:").pack(side=tk.LEFT, padx=5)
        self.filter_value = ttk.Entry(filter_frame)
        self.filter_value.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        ttk.Button(filter_frame, text="Apply", command=self.apply_filter).pack(side=tk.LEFT, padx=5)
        ttk.Button(filter_frame, text="Clear", command=self.clear_filter).pack(side=tk.LEFT, padx=5)
        
        # Treeview with scrollbars
        tree_frame = ttk.Frame(results_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        yscroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        xscroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL)
        
        self.results_tree = ttk.Treeview(tree_frame, 
                                      yscrollcommand=yscroll.set,
                                      xscrollcommand=xscroll.set)
        self.results_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        yscroll.config(command=self.results_tree.yview)
        xscroll.config(command=self.results_tree.xview)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        xscroll.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN)
        self.status_bar.pack(fill=tk.X)
        
        # Export and Settings buttons
        btn_frame = ttk.Frame(left_panel)
        btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(btn_frame, text="Export to PDF", 
                 command=self.export_to_pdf).pack(side=tk.LEFT, expand=True)
        ttk.Button(btn_frame, text="Export to CSV", 
                 command=self.export_to_csv).pack(side=tk.LEFT, expand=True)
        ttk.Button(btn_frame, text="Export to Excel", 
                 command=self.export_to_excel).pack(side=tk.LEFT, expand=True)
        ttk.Button(btn_frame, text="Settings", 
                 command=self.show_connection_dialog).pack(side=tk.LEFT, expand=True)

    def export_to_excel(self):
        selected_items = self.results_tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select rows to export")
            return
            
        try:
            self.status_var.set("Preparing Excel export...")
            self.root.update()
            
            from datetime import datetime
            import openpyxl
            
            # Get column names
            cols = self.results_tree["columns"]
            
            # Create filename with timestamp
            filename = f"airport_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Airport Data"
            
            # Write header
            for col_num, col in enumerate(cols, 1):
                ws.cell(row=1, column=col_num, value=col)
            
            # Write data rows
            for row_num, item in enumerate(selected_items, 2):
                values = self.results_tree.item(item)['values']
                for col_num, val in enumerate(values, 1):
                    ws.cell(row=row_num, column=col_num, value=val)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(filename)
            
            self.status_var.set(f"Excel exported to {filename}")
            messagebox.showinfo("Success", f"Excel exported successfully to {filename}")
            
        except ImportError:
            error_msg = "Please install openpyxl package (pip install openpyxl)"
            self.status_var.set(error_msg)
            messagebox.showerror("Error", error_msg)
        except Exception as e:
            error_msg = f"Failed to export Excel: {e}"
            self.status_var.set(error_msg)
            messagebox.showerror("Error", error_msg)

    def export_to_csv(self):
        selected_items = self.results_tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select rows to export")
            return
            
        try:
            self.status_var.set("Preparing CSV export...")
            self.root.update()
            
            from datetime import datetime
            import csv
            
            # Get column names
            cols = self.results_tree["columns"]
            
            # Create filename with timestamp
            filename = f"airport_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            with open(filename, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                writer.writerow(cols)
                
                # Write data rows
                for item in selected_items:
                    values = self.results_tree.item(item)['values']
                    writer.writerow(values)
            
            self.status_var.set(f"CSV exported to {filename}")
            messagebox.showinfo("Success", f"CSV exported successfully to {filename}")
            
        except Exception as e:
            error_msg = f"Failed to export CSV: {e}"
            self.status_var.set(error_msg)
            messagebox.showerror("Error", error_msg)

    def export_to_pdf(self):
        selected_items = self.results_tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select rows to export")
            return
            
        try:
            self.status_var.set("Preparing PDF export...")
            self.root.update()
            
            from fpdf import FPDF
            from datetime import datetime
            
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            
            # Add title with logo and timestamp
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(200, 10, txt="✈ Airport Database Export", ln=1, align='C')
            pdf.set_font("Arial", '', 10)
            pdf.cell(200, 10, txt=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1, align='C')
            
            # Get column names
            cols = self.results_tree["columns"]
            
            # Add table header
            pdf.set_fill_color(200, 220, 255)
            pdf.set_font("Arial", 'B', 12)
            for col in cols:
                pdf.cell(40, 10, txt=col, border=1, fill=True)
            pdf.ln()
            
            # Add data rows
            pdf.set_font("Arial", '', 10)
            for item in selected_items:
                values = self.results_tree.item(item)['values']
                for val in values:
                    pdf.cell(40, 10, txt=str(val), border=1)
                pdf.ln()
                
            # Save the PDF with timestamp
            filename = f"airport_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            pdf.output(filename)
            
            self.status_var.set(f"PDF exported to {filename}")
            messagebox.showinfo("Success", f"PDF exported successfully to {filename}")
            
        except ImportError:
            error_msg = "Please install fpdf package (pip install fpdf)"
            self.status_var.set(error_msg)
            messagebox.showerror("Error", error_msg)
        except Exception as e:
            error_msg = f"Failed to export PDF: {e}"
            self.status_var.set(error_msg)
            messagebox.showerror("Error", error_msg)

    # [Rest of the methods...]
    def connect_to_db(self):
        try:
            logger.debug(f"Attempting to connect to MySQL with config: {self.db_config}")
            self.connection = mysql.connector.connect(
                host=self.db_config['host'],
                database=self.db_config['database'],
                user=self.db_config['user'],
                password=self.db_config['password']
            )
            if self.connection.is_connected():
                logger.info(f"Successfully connected to {self.db_config['database']}")
                self.status_var.set(f"Connected to {self.db_config['database']}")
                return True
            logger.warning("Connection established but not connected")
            return False
        except Error as e:
            logger.error(f"Database connection failed: {e}", exc_info=True)
            messagebox.showerror("Error", f"Database connection failed: {e}")
            return False

    def apply_filter(self):
        filter_col = self.filter_column.get()
        filter_val = self.filter_value.get().lower()
        
        if not filter_col or not filter_val:
            return
            
        try:
            # Get all items
            all_items = self.results_tree.get_children()
            
            # Show only matching items
            for item in all_items:
                values = self.results_tree.item(item)['values']
                col_index = self.results_tree['columns'].index(filter_col)
                if str(values[col_index]).lower().find(filter_val) >= 0:
                    self.results_tree.item(item, tags=('visible',))
                else:
                    self.results_tree.item(item, tags=('hidden',))
                    
            self.results_tree.tag_configure('visible', foreground='black')
            self.results_tree.tag_configure('hidden', foreground='gray90')
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to apply filter: {e}")

    def clear_filter(self):
        self.filter_column.set('')
        self.filter_value.delete(0, tk.END)
        
        for item in self.results_tree.get_children():
            self.results_tree.item(item, tags=('visible',))
        self.results_tree.tag_configure('visible', foreground='black')

    def execute_query(self):
        query = self.query_entry.get("1.0", tk.END).strip()
        if not query:
            messagebox.showwarning("Warning", "No query entered")
            return
            
        try:
            if not self.connection or not self.connection.is_connected():
                if not self.connect_to_db():
                    raise Error("Not connected to database")
            
            cursor = self.connection.cursor()
            cursor.execute(query)
            
            # Clear previous results
            for item in self.results_tree.get_children():
                self.results_tree.delete(item)
                
            # Get column names
            columns = [desc[0] for desc in cursor.description]
            self.results_tree["columns"] = columns
            # Auto-size columns based on content
            for col in columns:
                self.results_tree.heading(col, text=col)
                # Set initial width based on column name length
                col_width = min(200, max(80, len(col) * 8))
                self.results_tree.column(col, width=col_width, minwidth=80, stretch=tk.YES)
            
            # Update columns after data is inserted
            def update_columns():
                for col in columns:
                    max_len = max([len(str(self.results_tree.set(child, col))) 
                                 for child in self.results_tree.get_children()] or [0])
                    new_width = min(300, max(80, max_len * 7))
                    self.results_tree.column(col, width=new_width)
                self.results_tree.update()
            
            self.root.after(100, update_columns)
                
            # Update filter dropdown with column names
            self.filter_column['values'] = columns
            
            # Add data rows
            for row in cursor:
                self.results_tree.insert("", tk.END, values=row, tags=('visible',))
                
            rowcount = cursor.rowcount
            self.status_var.set(f"Query executed - {rowcount} rows affected")
            messagebox.showinfo("Success", f"Query executed successfully\n{rowcount} rows affected")
            cursor.close()
            
        except Error as e:
            error_msg = f"Query failed: {str(e)}"
            self.status_var.set(error_msg)
            messagebox.showerror("Error", error_msg)
            print(f"Query Error: {error_msg}")  # Debug output

    def set_query(self, query):
        self.query_entry.delete("1.0", tk.END)
        self.query_entry.insert("1.0", query)

    def delete_selected(self):
        selected_items = self.results_tree.selection()
        if not selected_items:
            messagebox.showwarning("Warning", "Please select rows to delete")
            return
            
        # Confirm deletion
        confirm = messagebox.askyesno(
            "Confirm Delete",
            f"Are you sure you want to delete {len(selected_items)} record(s)?",
            icon='warning'
        )
        if not confirm:
            self.status_var.set("Delete operation cancelled")
            return
            
        try:
            self.status_var.set("Deleting records...")
            self.root.update()
            
            if not self.connection or not self.connection.is_connected():
                if not self.connect_to_db():
                    raise Error("Not connected to database")
            
            cursor = self.connection.cursor()
            
            # Get table name from the query
            query = self.query_entry.get("1.0", tk.END).strip().lower()
            if "from" in query:
                table_name = query.split("from")[1].split()[0]
            else:
                table_name = self.results_tree['columns'][0].lower()
            
            # Get primary key column
            cursor.execute(f"SHOW KEYS FROM {table_name} WHERE Key_name = 'PRIMARY'")
            pk_info = cursor.fetchone()
            if not pk_info:
                raise Error(f"No primary key found for table {table_name}")
            primary_key = pk_info[4]  # Column_name is at index 4
            
            deleted_count = 0
            for item in selected_items:
                try:
                    values = self.results_tree.item(item)['values']
                    # Find the index of the primary key in the displayed columns
                    pk_index = self.results_tree['columns'].index(primary_key)
                    pk_value = values[pk_index]
                    
                    query = f"DELETE FROM {table_name} WHERE {primary_key} = %s"
                    cursor.execute(query, (pk_value,))
                    deleted_count += 1
                except Error as e:
                    logger.error(f"Failed to delete record: {e}")
                    continue
            
            self.connection.commit()
            self.status_var.set(f"Successfully deleted {deleted_count} of {len(selected_items)} records")
            
            # Refresh the view
            self.execute_query()
            
        except Error as e:
            error_msg = f"Failed to delete records: {e}"
            logger.error(error_msg)
            self.status_var.set(error_msg)
            messagebox.showerror("Error", error_msg)

    def get_table_columns(self, table_name):
        """Get column names and types for a table"""
        try:
            cursor = self.connection.cursor()
            cursor.execute(f"DESCRIBE {table_name}")
            columns = []
            for (field, type, null, key, default, extra) in cursor:
                columns.append({
                    'name': field,
                    'type': type,
                    'required': null == 'NO' and not default and not extra == 'auto_increment'
                })
            return columns
        except Error as e:
            messagebox.showerror("Error", f"Failed to get table schema: {e}")
            return []

    def show_add_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Add New Record")
        dialog.geometry("500x400")
        
        # Main container
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Table selection
        table_frame = ttk.Frame(main_frame)
        table_frame.pack(fill=tk.X, pady=5)
        ttk.Label(table_frame, text="Table:").pack(side=tk.LEFT)
        table_var = tk.StringVar()
        table_cb = ttk.Combobox(table_frame, textvariable=table_var, 
                              values=["flight", "passenger", "schedule", "journey", "ticket"])
        table_cb.pack(side=tk.LEFT, expand=True, fill=tk.X)
        
        # Form fields container
        form_frame = ttk.Frame(main_frame)
        form_frame.pack(fill=tk.BOTH, expand=True)
        self.add_fields = {}
        
        def update_form_fields(*args):
            # Clear previous fields
            for widget in form_frame.winfo_children():
                widget.destroy()
            self.add_fields.clear()
            
            table = table_var.get()
            if not table:
                return
                
            columns = self.get_table_columns(table)
            if not columns:
                return
                
            # Create form fields
            for i, col in enumerate(columns):
                row = ttk.Frame(form_frame)
                row.pack(fill=tk.X, pady=2)
                
                label = f"{col['name']} ({col['type']})"
                if col['required']:
                    label += " *"
                ttk.Label(row, text=label, width=25).pack(side=tk.LEFT)
                
                entry = ttk.Entry(row)
                entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
                self.add_fields[col['name']] = entry
                
                # Add example text for guidance
                if "int" in col['type']:
                    entry.insert(0, "123")
                elif "date" in col['type'] or "time" in col['type']:
                    entry.insert(0, "YYYY-MM-DD" if "date" in col['type'] else "HH:MM:SS")
                elif "varchar" in col['type']:
                    entry.insert(0, "text value")
        
        table_var.trace_add("write", update_form_fields)
        
        # Submit button
        def submit_record():
            try:
                table = table_var.get()
                if not table:
                    raise ValueError("Please select a table")
                    
                data = {}
                for col in self.get_table_columns(table):
                    value = self.add_fields[col['name']].get()
                    if col['required'] and not value:
                        raise ValueError(f"{col['name']} is required")
                    data[col['name']] = value
                
                if not self.connection or not self.connection.is_connected():
                    if not self.connect_to_db():
                        raise Error("Not connected to database")
                
                cursor = self.connection.cursor()
                columns = ", ".join(data.keys())
                values = ", ".join([f"'{v}'" for v in data.values()])
                query = f"INSERT INTO {table} ({columns}) VALUES ({values})"
                cursor.execute(query)
                self.connection.commit()
                messagebox.showinfo("Success", "Record added successfully")
                dialog.destroy()
                
                # Refresh the view
                self.execute_query()
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add record: {e}")

        ttk.Button(dialog, text="Submit", command=submit_record).pack(pady=10)

    def show_connection_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("Database Connection Settings")
        dialog.geometry("300x200")
        
        ttk.Label(dialog, text="Database Configuration").pack(pady=10)
        
        fields = {}
        for field in ['host', 'user', 'password', 'database']:
            frame = ttk.Frame(dialog)
            frame.pack(fill=tk.X, padx=5, pady=5)
            ttk.Label(frame, text=field.capitalize()+":", width=10).pack(side=tk.LEFT)
            entry = ttk.Entry(frame)
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            fields[field] = entry
            entry.insert(0, self.db_config.get(field, ''))
            
        def save_settings():
            for field, entry in fields.items():
                self.db_config[field] = entry.get()
            dialog.destroy()
            self.connect_to_db()
            
        ttk.Button(dialog, text="Save", command=save_settings).pack(pady=10)


if __name__ == "__main__":
    root = tk.Tk()
    app = AirportDBGUI(root)
    root.mainloop()
